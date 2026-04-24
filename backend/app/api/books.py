"""REST endpoints for the stepper-driven book flow.

Each LLM-backed action is a synchronous request — callers see a spinner while
the mutation is in-flight, then the fresh resource on return. No background
graph runs, no interrupt/resume plumbing.
"""

import logging
from io import BytesIO
from typing import Any, Literal, Optional

from fastapi import APIRouter, File, HTTPException, Response, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict

from app.db import repositories as repo
from app.services import chapter_service, compile_service, outline_service

MAX_BULK_ROWS = 50

log = logging.getLogger(__name__)

router = APIRouter(prefix="/books", tags=["books"])


# ---------- schemas ----------

class CreateBookPayload(BaseModel):
    title: str


class BookResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    title: str
    status: str
    current_node: Optional[str] = None
    final_docx_path: Optional[str] = None
    final_txt_path: Optional[str] = None
    final_pdf_path: Optional[str] = None
    created_at: str
    updated_at: str


class OutlineResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    book_id: str
    version: int
    parent_id: Optional[str] = None
    content: dict[str, Any]
    status: str
    created_at: str


class ChapterResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")

    id: str
    book_id: str
    index: int
    version: int
    parent_id: Optional[str] = None
    title: Optional[str] = None
    content_md: str
    summary: Optional[str] = None
    status: str
    created_at: str


class RevisePayload(BaseModel):
    note: Optional[str] = None


class EditOutlinePayload(BaseModel):
    content: dict[str, Any]


class DownloadResponse(BaseModel):
    url: str
    format: str
    path: str


DownloadFormat = Literal["docx", "pdf", "txt", "md"]


# ---------- helpers ----------

def _require_book(book_id: str) -> dict[str, Any]:
    book = repo.get_book(book_id)
    if not book:
        raise HTTPException(status_code=404, detail="book not found")
    return book


def _signed(path: str) -> str:
    url = compile_service.create_signed_url(path, expires_in_seconds=3600)
    if not url:
        raise HTTPException(status_code=500, detail="could not sign storage url")
    return url


# ---------- books ----------

@router.get("", response_model=list[BookResponse])
def list_books() -> list[BookResponse]:
    return [BookResponse(**b) for b in repo.list_books()]


@router.post("", response_model=BookResponse, status_code=201)
def create_book(payload: CreateBookPayload) -> BookResponse:
    """Create a book. Outline generation is driven by a separate CTA."""
    book = repo.create_book(payload.title)
    return BookResponse(**book)


@router.post("/bulk", response_model=list[BookResponse], status_code=201)
async def create_books_bulk(file: UploadFile = File(...)) -> list[BookResponse]:
    """Bulk-create books from an .xlsx file.

    Format: column A holds titles, row 1 is a header (ignored), data starts
    at row 2. Blank cells are skipped. Capped at `MAX_BULK_ROWS`.
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="expected a .xlsx file")

    content = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"invalid .xlsx: {exc}") from exc

    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="workbook has no sheets")

    titles: list[str] = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        value = row[0]
        if isinstance(value, str) and value.strip():
            titles.append(value.strip())
        if len(titles) >= MAX_BULK_ROWS:
            break

    if not titles:
        raise HTTPException(
            status_code=400,
            detail="no titles found — put titles in column A starting at row 2",
        )

    created = [BookResponse(**repo.create_book(t)) for t in titles]
    return created


@router.get("/{book_id}", response_model=BookResponse)
def get_book(book_id: str) -> BookResponse:
    return BookResponse(**_require_book(book_id))


@router.post("/{book_id}/restart", response_model=BookResponse)
def restart_book(book_id: str) -> BookResponse:
    """Wipe outlines/chapters/feedback/storage and reset the book to `created`.

    Used from the failure-recovery card; keeps id/title/created_at stable so
    the UI retains its identity.
    """
    _require_book(book_id)
    repo.delete_book_storage(book_id)
    repo.reset_book(book_id)
    refreshed = repo.get_book(book_id)
    assert refreshed is not None
    return BookResponse(**refreshed)


@router.delete("/{book_id}", status_code=204, response_class=Response)
def delete_book(book_id: str) -> Response:
    _require_book(book_id)
    repo.delete_book(book_id)
    repo.delete_book_storage(book_id)
    return Response(status_code=204)


# ---------- outlines ----------

@router.get("/{book_id}/outlines", response_model=list[OutlineResponse])
def list_outlines(book_id: str) -> list[OutlineResponse]:
    _require_book(book_id)
    return [OutlineResponse(**o) for o in repo.list_outlines(book_id)]


@router.post("/{book_id}/outline/generate", response_model=OutlineResponse)
def generate_outline(book_id: str) -> OutlineResponse:
    _require_book(book_id)
    if repo.get_latest_outline(book_id):
        raise HTTPException(
            status_code=409,
            detail="outline already exists — use revise to regenerate",
        )
    try:
        row = outline_service.generate(book_id)
    except Exception as exc:  # noqa: BLE001
        log.exception("outline generation failed for %s", book_id)
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return OutlineResponse(**row)


@router.post("/{book_id}/outline/revise", response_model=OutlineResponse)
def revise_outline(book_id: str, payload: RevisePayload) -> OutlineResponse:
    _require_book(book_id)
    latest = repo.get_latest_outline(book_id)
    if not latest:
        raise HTTPException(status_code=400, detail="no outline to revise")
    if latest["status"] == "approved":
        raise HTTPException(status_code=409, detail="outline is already approved")
    try:
        row = outline_service.revise(book_id, payload.note)
    except Exception as exc:  # noqa: BLE001
        log.exception("outline revision failed for %s", book_id)
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return OutlineResponse(**row)


@router.post("/{book_id}/outline/edit", response_model=OutlineResponse)
def edit_outline(book_id: str, payload: EditOutlinePayload) -> OutlineResponse:
    _require_book(book_id)
    try:
        row = outline_service.edit(book_id, payload.content)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return OutlineResponse(**row)


@router.post("/{book_id}/outline/approve", response_model=OutlineResponse)
def approve_outline(book_id: str) -> OutlineResponse:
    _require_book(book_id)
    try:
        row = outline_service.approve(book_id)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return OutlineResponse(**row)


# ---------- chapters ----------

@router.get("/{book_id}/chapters", response_model=list[ChapterResponse])
def list_chapters(book_id: str) -> list[ChapterResponse]:
    _require_book(book_id)
    return [ChapterResponse(**c) for c in repo.list_chapters(book_id)]


@router.post("/{book_id}/chapters/draft", response_model=list[ChapterResponse])
def draft_chapter_slots(book_id: str) -> list[ChapterResponse]:
    """Create one empty slot per outline chapter. Idempotent."""
    _require_book(book_id)
    try:
        rows = chapter_service.create_slots(book_id)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return [ChapterResponse(**c) for c in rows]


@router.post(
    "/{book_id}/chapters/{index}/generate", response_model=ChapterResponse
)
def generate_chapter(book_id: str, index: int) -> ChapterResponse:
    _require_book(book_id)
    try:
        row = chapter_service.generate(book_id, index)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        log.exception("chapter generation failed for %s/%s", book_id, index)
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return ChapterResponse(**row)


@router.post(
    "/{book_id}/chapters/{index}/revise", response_model=ChapterResponse
)
def revise_chapter(
    book_id: str, index: int, payload: RevisePayload
) -> ChapterResponse:
    _require_book(book_id)
    try:
        row = chapter_service.revise(book_id, index, payload.note)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        log.exception("chapter revision failed for %s/%s", book_id, index)
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return ChapterResponse(**row)


@router.post(
    "/{book_id}/chapters/{index}/approve", response_model=ChapterResponse
)
def approve_chapter(book_id: str, index: int) -> ChapterResponse:
    _require_book(book_id)
    try:
        row = chapter_service.approve(book_id, index)
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        log.exception("chapter approval failed for %s/%s", book_id, index)
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return ChapterResponse(**row)


# ---------- downloads ----------

@router.get("/{book_id}/download", response_model=DownloadResponse)
def download_combined(
    book_id: str, format: DownloadFormat = "pdf"
) -> DownloadResponse:
    """Compile outline + approved chapters into a single file."""
    _require_book(book_id)
    try:
        path = compile_service.compile_combined(book_id, format)
    except RuntimeError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return DownloadResponse(url=_signed(path), format=format, path=path)


@router.get(
    "/{book_id}/chapters/{index}/download", response_model=DownloadResponse
)
def download_chapter(
    book_id: str, index: int, format: DownloadFormat = "pdf"
) -> DownloadResponse:
    _require_book(book_id)
    try:
        path = compile_service.compile_chapter(book_id, index, format)
    except RuntimeError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return DownloadResponse(url=_signed(path), format=format, path=path)
